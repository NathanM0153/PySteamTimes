#read from sheet, copy to other sheets in categories
import openpyxl
import sys

file = "SteamGames.xlsx"
wb = openpyxl.load_workbook(file)

def readSheet():
    games = []
    ws = wb["Sheet"]
    for row in ws.iter_rows():
        for cell in row:
            games.append(cell.value)
    return games


def getSheet():
    try:
        x = input("\nSelect sheet:\n"
                  "1. Play Soon\n"
                  "2. Play Someday\n"
                  "3. Play Never\n" 
                  "4. Finished\n"
                  "5. Dropped or Unfinished\n\n")
    except:
        print("Please close the file before making changes.")
        sys.exit()

    if x == "1":
        return "Soon"
    elif x == "2":
        return "Someday"
    elif x == "3":
        return "Never"
    elif x == "4":
        return "Finished"
    elif x == "5":
        return "Unfinished"
    elif x.upper() == "QUIT":
        wb.save(file)
        print("File saved.")
        return None
    else:
        return "Other"


def exportToExcel(gameList):
    if len(wb.sheetnames) == 1: #prevents sheets being made multiple times on subsequent runs
        ws1 = wb.create_sheet("Soon")
        ws2 = wb.create_sheet("Someday")
        ws3 = wb.create_sheet("Never")
        ws4 = wb.create_sheet("Finished")
        ws5 = wb.create_sheet("Unfinished")
        ws6 = wb.create_sheet("Other")
    gameSheet = wb.active
    x = input("Start from?\n")
    x = int(x)  
    for i in range(x,len(gameList)):
        print(gameList[i])
        sheet = getSheet()
        if sheet == None:
            print("Games finished:", i)
            sys.exit()
        gameSheet = wb[sheet]
        gameSheet.cell(row=i+1, column=1, value=gameList[i])
        wb.save(file)

class separateGames():

    def main():
        print("Enter 'quit' to save and quit.")
        games = readSheet()
        for i in range(0,len(games)):   
            exportToExcel(games)
        return 0
    
    if __name__ == "__main__":
        main()
