import sys
import openpyxl
import returnQuery #in working directory

            

para = returnQuery.loadWorkbook() #returns list of parameters
wb = para[0]
ws = para[1]
fileName = para[2]
wsList = wb.sheetnames
gameCount = ws.max_row
typeList = []
searchResult = [] 

def exportToSheet(timeList, typeList):
    y = input("Is this your first time running through this script? Y/N\n")
    if y == "N" or y == "n":
        timeColumn = input("Select the Excel column you would like to put the times in. If you would like to enter in column C, please enter 3, etc. "
                          "Please keep in mind that the game list is in column A and the first set of times created was put in column B.\n")
        if timeColumn == "1":
            z = input("Are you sure? This will overwrite your games. Y/N\n")
            if z == "N" or z == "n":
                exportToSheet(timeList, typeList)
            elif z != "Y" and z != "y":
                print("Invalid input.")
                exportToSheet(timeList, typeList)
        elif timeColumn == "2":
            z = input("Are you sure? This will overwrite your times in column B. Y/N\n")
            if z == "N" or z == "n":
                exportToSheet(timeList, typeList)
            elif z != "Y" and z != "y":
                print("Invalid input.")
                exportToSheet(timeList, typeList)
                
        #if not first time, they already answered these questions, no need to ask again
        multi = "N"
        endless = "N"
    elif y == "Y" or y == "y":
        #defaults to column B
        timeColumn = "2"
        #multi = input("\nWould you like to remove primarily multiplayer games? These are often given inflated completion times. Y/N\n")
        #endless = input("Would you like to remove games marked as endless? These are often given arbitrary completion times. Y/N\n")
    else:
        print("Invalid input.")
        exportToSheet(timeList, typeList)
    #check if column input is integer
    try:
        timeColumn = int(timeColumn)
        if timeColumn < 1:
            print("Invalid column input.")
            exportToSheet(timeList, typeList) 
    except:
        print("Invalid column input.")
        exportToSheet(timeList, typeList)

    try:
        insertTime(typeList, timeList, timeColumn)
        #insertType(typeList, timelist, timeColumn)
        #delMulti(typeList, multi)
        #delEndless(typeList, endless)
        print("Exported to file successfully.")
        print("Refer to the README.txt in order to complete results.")
    except Exception as e:
        print("Export not successful. Refer to the README for instructions.")
        print(e)
        sys.exit()


def insertTime(typeList, times, col):
    for i in range(0, len(times)):
        gameTime = ws.cell(row=i+1, column=col)
        gameTime.value = times[i]
    wb.save(fileName)


# these currently unused
def insertType(typeList, times, col):
    for i in range(0, len(times)):
        gameType = ws.cell(row=i+1, column=col+1)
        try:
            gameType.value = typeList[i]
        except:
            gameType.value = ""
    wb.save(fileName)
    
def delMulti(typeList, multi):
    if (multi == "Y" or multi == "y"):
        #delete rows bottom to top or the indexes get screwed up
        for i in reversed(range(0,gameCount)):
            if typeList[i] == "multi":
                ws.delete_rows(i+1,1)
    elif (multi != "N" and multi != "n"):
        print("Invalid input multi.")
        delMulti(typeList, multi)
    else:
        pass
    wb.save(fileName)

def delEndless(typeList, endless):
    if (endless == "Y" or endless == "y"):     
        for i in reversed(range(0,gameCount)):
            if typeList[i] == "endless":
                ws.delete_rows(i+1,1)
    elif (endless != "N" and endless != "n"):
        print("Invalid input endless.")
        delEndless(typeList, multi)
    wb.save(fileName)


class HLTBSteam():
    def main():
        first_col = list(ws['A'])
        gameList = returnQuery.getGames(first_col)
        result = returnQuery.getTimes(gameList) #returns nested list
        timeList = result[0]
        typeList = result[1]
        exportToSheet(timeList, typeList)
        return 0

    if __name__ == "__main__":
        main()
