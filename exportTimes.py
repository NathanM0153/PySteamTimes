# modules needed: aiohttp, requests, fake_useragent, openpyxl, xlwt, steamfront

from howlongtobeatpy import HowLongToBeat
# https://pypi.org/project/howlongtobeatpy/

import xlwt
from xlwt import Workbook
# https://www.geeksforgeeks.org/writing-excel-sheet-using-python/
from openpyxl import load_workbook
# https://ehmatthes.github.io/pcc_2e/beyond_pcc/extracting_from_excel/

filePath = "C:/Users/Nathan/Documents/PySteamTimer/SteamGames.xlsx"
try:
    wb = load_workbook(filePath)
    ws = wb["run_results"] #raw input in main?
except:
    print("A problem occurred. Here's how to fix it:")
    print("Open Excel, go to File -> Open -> Browse.")
    print("Single-click on SteamGames.xlsx, and navigate down to the Open button on the bottom right.")
    print("Click the down arrow next to Open, click Open and Repair, Extract Data, Convert to Values.")
    print("FInally, go back to File -> Save As, and save the file as SteamGames.xlsx.")
    print("You may have to change the drop down below the file name from .xls to .xlsx.")
    print("After that, you can rerun this script.")
    quit(0)

gameCount = ws.max_row

# takes in JSON, returns completion time as int
def completion_time(game):
    if game is not None:
        time = (game.main_story + game.main_extra) // 2
        return time
    else:
        return 0

def getGames(): 
    gameList = []
    rows = list(ws.rows)
    for i in range(0,gameCount):
        for j in rows[i]:
            gameList.append(j.value)
    return gameList

# takes in list of games, returns list of times for the games
def getTimes(games):
    timeList = []
    for i in range(0, gameCount):
        gameSearch = searchforGame(games[i])
        # print(gameSearch)
        timeList.append(completion_time(gameSearch))
    return timeList

def exportToSheet(times):
    for i in range(0, len(times)):
        ws.cell(row=i + 1, column=2).value = times[i]
        # ws.write(i,1,times[i])
    wb.save('SteamGameTimes_Finished.xlsx')
    print("Exported to file successfully.")

# takes in a string, returns a JSON result to be interpreted
def searchforGame(name):
    results = HowLongToBeat().search(name, similarity_case_sensitive=False)
    best_element = ""
    if results is not None and len(results) > 0:
        best_element = max(results, key=lambda element: element.similarity)
        print(name, "found")
        return best_element
    else:
        print(name,  "NOT found. Check the name listed in your spreadsheet and ensure it does not include extraneous info such as Definitive Edition.")
        return None

# takes in JSON, returns name of game as string
def getName(game):
    return game.game_name

class HLTBSteam():
    def main():
        gameList = getGames()
        timeList = getTimes(gameList)
        for i in range(0,gameCount):
            print(gameList[i], end=": ")
            print(timeList[i], "hours")
        exportToSheet(timeList)
        return 0

    if __name__ == "__main__":
        main()
