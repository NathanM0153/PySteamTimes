import sys
import openpyxl
import os
import pathlib


from howlongtobeatpy import HowLongToBeat
#credit: https://github.com/ScrappyCocco/HowLongToBeat-PythonAPI


def findFile():
    excelFilesinDir = []
    cwd = os.getcwd()
    excelFilesinDir = list(pathlib.Path(cwd).glob('*.xlsx'))
    #print(excelFilesinDir)
    if len(excelFilesinDir) == 1: #only one file
        return excelFilesinDir[0]
    elif len(excelFilesinDir) == 0: #no xlsx files, means user hasn't imported
        print("Please run importSteam.py first.")
        sys.exit()
    else:
        print("Multiple .xlsx files detected:")
        for i in range(0,len(excelFilesinDir)):
            print(i+1, " ", excelFilesinDir[i])
        x = input("Enter the listed number of the one you would like to use.\n")
        try:
            x = int(x)
            if x < 1:
                print("Invalid input.")
                findFile()
            print(excelFilesinDir[x-1], "selected.\n")
            global fileName
            fileName = excelFilesinDir[x-1]
        except Exception as e:
            print("Invalid input.")
            #print(e)
            findFile()
    #print(fileName)
    return fileName

def chooseSheet(wb):
    sheets = wb.sheetnames
    if len(sheets) == 1:
        return wb.sheetnames[0]
    else:
        print("Multiple sheets detected in file:")
        for i in range(0,len(sheets)):
            print(i+1, " ", sheets[i])
        x = input("Enter the listed number of the one you would like to use.\n")
        try:
            x = int(x)
            if x < 1:
                print("Invalid input.")
                chooseSheet(wb)
            elif x > len(sheets):
                print("Invalid input.")
                chooseSheet(wb)
            print(sheets[x-1], "selected.\n")
        except:
            print("Invalid input (error).")
            chooseSheet(wb)
    return sheets[x-1]

def loadWorkbook():
    fileName = findFile()
    wb = openpyxl.load_workbook(fileName)
    sheet = chooseSheet(wb)
    ws = wb[sheet]
    #wsList = wb.sheetnames
    #gameCount = ws.max_row
    #except Exception as e:
    #    print("A problem occurred:", e)
        #print("If your error is not apparent, please check the README.txt file for assistance.")
    #    sys.exit()
    return [wb, ws, fileName]

# pulls games from the excel sheet, puts in a list
def getGames(game_col): 
    gameList = []
    #try and get first_col in here instead of the whole data file
    #print(rows)
    try:
        for i in game_col:
            gameList.append(i.value)
    except Exception as e:
        print(e)
        print("Game count invalid. Check if your sheet is empty.")
        sys.exit()
    return gameList

# takes in JSON, returns completion time as int
def completion_time(game, playstyle):
    time = 0
    if game is not None:
        try:
            if playstyle == 1:
                time = round(game.main_story, 2)
            elif playstyle == 2:
                time = round(game.main_extra, 2)
            elif playstyle == 3:
                time = round(game.completionist, 2)
            elif playstyle == 4:
                time = round((game.main_story + game.main_extra) / 2, 2)
            elif playstyle == 5:
                time = round((game.main_extra + game.completionist) / 2, 2)
            elif playstyle == 6:
                time = round((game.main_story + game.main_extra + game.completionist) / 3, 2)
            else:
                print("Time not found")
        except Exception as e:
            print(e)
            sys.exit()
        return time
    else:
        return "Not found in HLTB database."

# takes in list of games, returns list of times for the games
def getTimes(games):
    x = input("Select the type of times you would like to retrieve:\n"
              "1. Main story completion only\n"
              "2. Main story + extra content\n"
              "3. Full completion\n"
              "4. Average of 1 and 2 (recommended)\n"
              "5. Average of 2 and 3\n"
              "6. Average of all\n\n"
              "If you would like multiple of the above, please re-run the script at the end.\n")
    print("")
    try:
        x = int(x) #this only works if x is an int, filters input
        if x < 1 or x > 6:
            print("Invalid input.")
            getTimes(games)            
    except:
        print("Invalid input.")
        getTimes(games)

    result = searchResults(games, x)

    return result

def searchResults(games, x):
    timeList = []
    typeList = []
    gameCount = len(games)
    for i in range(0, gameCount):
        gameSearch = searchforGame(games[i])
        #print(gameSearch)
        timeList.append(completion_time(gameSearch, x))
        try:
            typeList.append(gameSearch.game_type)
        except:
            typeList.append("")
    return [timeList, typeList]



# takes in a string, returns a JSON result to be interpreted
def searchforGame(name):
    try:
        results = HowLongToBeat().search(name, similarity_case_sensitive=False)
    except:
        print("Internet connection required. Please check your "
              "settings to ensure you have internet access.")
        sys.exit()
    best_element = ""
    if results is not None and len(results) > 0:
        best_element = max(results, key=lambda element: element.similarity)
        print(name, "found")
        #count += 1
        #maybe someday put a number here
        return best_element
    else:
        print(name,  "NOT found. Check the README file for reasons this may be.")
        return None
